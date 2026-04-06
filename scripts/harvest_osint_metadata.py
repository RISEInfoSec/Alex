#!/usr/bin/env python3
"""
harvest_osint_metadata.py

Batch-harvest bibliographic metadata for the OSINT + Cybersecurity Research Library.

What it does
------------
Reads an Excel workbook or CSV containing paper records and attempts to resolve,
from authoritative internet sources where available:

- DOI
- Authors
- Venue
- Abstract

Primary sources
---------------
1. Crossref API
2. OpenAlex API

Optional sources
----------------
3. Semantic Scholar API (title/abstract fallback only, if available)

Designed for batch processing in increments of 25, with caching and resumability.

Typical usage
-------------
python harvest_osint_metadata.py \
  --input osint_cyber_paper_library_206_v7_enriched.xlsx \
  --sheet Cyber_Corpus \
  --start-id 1 \
  --end-id 25 \
  --output osint_cyber_paper_library_batch1_harvested.xlsx \
  --mailto your-email@example.org

Requirements
------------
pip install pandas openpyxl requests

Notes
-----
- Crossref often does NOT expose full abstracts for every record.
- OpenAlex may expose abstract_inverted_index for some records.
- The script clearly marks provenance and whether an abstract is authoritative
  or a fallback.
- This script does not fabricate abstracts.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests


CROSSREF_WORKS = "https://api.crossref.org/works"
OPENALEX_WORKS = "https://api.openalex.org/works"
SEMANTIC_SCHOLAR_PAPER_SEARCH = "https://api.semanticscholar.org/graph/v1/paper/search"

DEFAULT_BATCH_SIZE = 25
REQUEST_TIMEOUT = 30
SLEEP_SECONDS = 0.8


# ----------------------------
# Utility helpers
# ----------------------------

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_title(title: str) -> str:
    t = clean_text(title).lower()
    t = re.sub(r"\barxiv\b", " ", t)
    t = re.sub(r"[^a-z0-9]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def similarity_key(title: str) -> str:
    words = normalize_title(title).split()
    return " ".join(words[:10])


def strip_jats_tags(text: str) -> str:
    if not text:
        return ""
    text = unescape(text)
    text = re.sub(r"</?jats:[^>]+>", " ", text)
    text = re.sub(r"</?[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def reconstruct_openalex_abstract(inverted_index: Optional[Dict[str, List[int]]]) -> str:
    if not inverted_index:
        return ""
    max_pos = -1
    for positions in inverted_index.values():
        if positions:
            max_pos = max(max_pos, max(positions))
    if max_pos < 0:
        return ""
    tokens = [""] * (max_pos + 1)
    for word, positions in inverted_index.items():
        for pos in positions:
            if 0 <= pos < len(tokens):
                tokens[pos] = word
    text = " ".join(tokens)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def unique_preserve_order(values: Iterable[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for v in values:
        v = clean_text(v)
        if not v:
            continue
        key = v.lower()
        if key not in seen:
            out.append(v)
            seen.add(key)
    return out


def format_authors_crossref(authors: List[Dict[str, Any]]) -> str:
    names = []
    for author in authors or []:
        given = clean_text(author.get("given"))
        family = clean_text(author.get("family"))
        name = clean_text(author.get("name"))
        if given or family:
            names.append(clean_text(f"{given} {family}"))
        elif name:
            names.append(name)
    return "; ".join(unique_preserve_order(names))


def format_authors_openalex(authorships: List[Dict[str, Any]]) -> str:
    names = []
    for authorship in authorships or []:
        author = authorship.get("author") or {}
        display_name = clean_text(author.get("display_name"))
        if display_name:
            names.append(display_name)
    return "; ".join(unique_preserve_order(names))


def venue_from_crossref(message: Dict[str, Any]) -> str:
    containers = message.get("container-title") or []
    if containers:
        return clean_text(containers[0])
    event = message.get("event") or {}
    if event.get("name"):
        return clean_text(event["name"])
    publisher = clean_text(message.get("publisher"))
    return publisher


def venue_from_openalex(work: Dict[str, Any]) -> str:
    primary = work.get("primary_location") or {}
    source = primary.get("source") or {}
    venue = clean_text(source.get("display_name"))
    if venue:
        return venue
    host = work.get("host_venue") or {}
    venue = clean_text(host.get("display_name"))
    if venue:
        return venue
    return ""


def choose_better(current: str, candidate: str) -> str:
    current = clean_text(current)
    candidate = clean_text(candidate)
    if not candidate:
        return current
    if not current:
        return candidate
    return candidate if len(candidate) > len(current) else current


# ----------------------------
# HTTP client with cache
# ----------------------------

class CachedHTTPClient:
    def __init__(self, cache_path: Path, mailto: Optional[str] = None):
        self.cache_path = cache_path
        self.mailto = mailto
        self.session = requests.Session()
        ua = "RISE-OSINT-Metadata-Harvester/1.0"
        if mailto:
            ua += f" (mailto:{mailto})"
        self.session.headers.update({
            "User-Agent": ua,
            "Accept": "application/json",
        })
        if cache_path.exists():
            try:
                self.cache = json.loads(cache_path.read_text(encoding="utf-8"))
            except Exception:
                self.cache = {}
        else:
            self.cache = {}

    def save_cache(self) -> None:
        self.cache_path.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

    def get_json(self, url: str, params: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
        cache_key = json.dumps({"url": url, "params": params or {}}, sort_keys=True)
        if cache_key in self.cache:
            return self.cache[cache_key]

        try:
            response = self.session.get(url, params=params, timeout=REQUEST_TIMEOUT)
            if response.status_code == 200:
                data = response.json()
                self.cache[cache_key] = data
                self.save_cache()
                time.sleep(SLEEP_SECONDS)
                return data
            else:
                self.cache[cache_key] = None
                self.save_cache()
                time.sleep(SLEEP_SECONDS)
                return None
        except Exception:
            self.cache[cache_key] = None
            self.save_cache()
            time.sleep(SLEEP_SECONDS)
            return None


# ----------------------------
# Match result model
# ----------------------------

@dataclass
class HarvestResult:
    doi: str = ""
    authors: str = ""
    venue: str = ""
    abstract: str = ""
    authority: str = ""
    abstract_status: str = ""
    source_url: str = ""
    score: float = 0.0


# ----------------------------
# Crossref / OpenAlex lookups
# ----------------------------

def crossref_by_doi(client: CachedHTTPClient, doi: str) -> Optional[HarvestResult]:
    doi = clean_text(doi)
    if not doi:
        return None
    url = f"{CROSSREF_WORKS}/{doi}"
    data = client.get_json(url)
    if not data or "message" not in data:
        return None
    message = data["message"]
    abstract = strip_jats_tags(message.get("abstract", ""))
    return HarvestResult(
        doi=clean_text(message.get("DOI", doi)),
        authors=format_authors_crossref(message.get("author", [])),
        venue=venue_from_crossref(message),
        abstract=abstract,
        authority="Crossref",
        abstract_status="publisher-native abstract" if abstract else "",
        source_url=clean_text(message.get("URL")),
        score=1.0,
    )


def crossref_search_title(client: CachedHTTPClient, title: str) -> Optional[HarvestResult]:
    title = clean_text(title)
    if not title:
        return None
    data = client.get_json(CROSSREF_WORKS, params={
        "query.title": title,
        "rows": 5,
        "select": "DOI,title,author,container-title,publisher,abstract,URL,event",
    })
    if not data:
        return None
    items = ((data.get("message") or {}).get("items") or [])
    if not items:
        return None

    target_norm = normalize_title(title)
    best: Optional[HarvestResult] = None
    best_score = -1.0

    for item in items:
        titles = item.get("title") or []
        candidate_title = clean_text(titles[0] if titles else "")
        if not candidate_title:
            continue
        cand_norm = normalize_title(candidate_title)
        score = 1.0 if cand_norm == target_norm else 0.0
        if similarity_key(candidate_title) == similarity_key(title):
            score = max(score, 0.85)
        elif target_norm in cand_norm or cand_norm in target_norm:
            score = max(score, 0.7)

        if score > best_score:
            abstract = strip_jats_tags(item.get("abstract", ""))
            best = HarvestResult(
                doi=clean_text(item.get("DOI")),
                authors=format_authors_crossref(item.get("author", [])),
                venue=venue_from_crossref(item),
                abstract=abstract,
                authority="Crossref search",
                abstract_status="publisher-native abstract" if abstract else "",
                source_url=clean_text(item.get("URL")),
                score=score,
            )
            best_score = score

    return best if best and best.score >= 0.7 else None


def openalex_by_doi(client: CachedHTTPClient, doi: str, mailto: Optional[str]) -> Optional[HarvestResult]:
    doi = clean_text(doi)
    if not doi:
        return None
    doi_url = doi if doi.lower().startswith("https://doi.org/") else f"https://doi.org/{doi}"
    params = {"filter": f"doi:{doi_url}"}
    if mailto:
        params["mailto"] = mailto
    data = client.get_json(OPENALEX_WORKS, params=params)
    if not data:
        return None
    results = data.get("results") or []
    if not results:
        return None
    work = results[0]
    abstract = reconstruct_openalex_abstract(work.get("abstract_inverted_index"))
    return HarvestResult(
        doi=clean_text((work.get("ids") or {}).get("doi", "")).replace("https://doi.org/", ""),
        authors=format_authors_openalex(work.get("authorships", [])),
        venue=venue_from_openalex(work),
        abstract=abstract,
        authority="OpenAlex",
        abstract_status="OpenAlex abstract" if abstract else "",
        source_url=clean_text((work.get("primary_location") or {}).get("landing_page_url", "")),
        score=1.0,
    )


def openalex_search_title(client: CachedHTTPClient, title: str, mailto: Optional[str]) -> Optional[HarvestResult]:
    title = clean_text(title)
    if not title:
        return None
    params = {"search": title, "per-page": 5}
    if mailto:
        params["mailto"] = mailto
    data = client.get_json(OPENALEX_WORKS, params=params)
    if not data:
        return None
    results = data.get("results") or []
    if not results:
        return None

    target_norm = normalize_title(title)
    best: Optional[HarvestResult] = None
    best_score = -1.0

    for work in results:
        candidate_title = clean_text(work.get("title"))
        if not candidate_title:
            continue
        cand_norm = normalize_title(candidate_title)
        score = 1.0 if cand_norm == target_norm else 0.0
        if similarity_key(candidate_title) == similarity_key(title):
            score = max(score, 0.85)
        elif target_norm in cand_norm or cand_norm in target_norm:
            score = max(score, 0.7)

        if score > best_score:
            abstract = reconstruct_openalex_abstract(work.get("abstract_inverted_index"))
            doi = clean_text((work.get("ids") or {}).get("doi", "")).replace("https://doi.org/", "")
            best = HarvestResult(
                doi=doi,
                authors=format_authors_openalex(work.get("authorships", [])),
                venue=venue_from_openalex(work),
                abstract=abstract,
                authority="OpenAlex search",
                abstract_status="OpenAlex abstract" if abstract else "",
                source_url=clean_text((work.get("primary_location") or {}).get("landing_page_url", "")),
                score=score,
            )
            best_score = score

    return best if best and best.score >= 0.7 else None


def semantic_scholar_search_title(client: CachedHTTPClient, title: str) -> Optional[HarvestResult]:
    title = clean_text(title)
    if not title:
        return None
    data = client.get_json(SEMANTIC_SCHOLAR_PAPER_SEARCH, params={
        "query": title,
        "limit": 5,
        "fields": "title,abstract,authors,venue,year,externalIds,url",
    })
    if not data:
        return None
    results = data.get("data") or []
    if not results:
        return None

    target_norm = normalize_title(title)
    best = None
    best_score = -1.0

    for item in results:
        candidate_title = clean_text(item.get("title"))
        if not candidate_title:
            continue
        cand_norm = normalize_title(candidate_title)
        score = 1.0 if cand_norm == target_norm else 0.0
        if similarity_key(candidate_title) == similarity_key(title):
            score = max(score, 0.85)
        elif target_norm in cand_norm or cand_norm in target_norm:
            score = max(score, 0.7)

        if score > best_score:
            authors = "; ".join(unique_preserve_order(
                clean_text((a or {}).get("name")) for a in (item.get("authors") or [])
            ))
            external_ids = item.get("externalIds") or {}
            doi = clean_text(external_ids.get("DOI"))
            best = HarvestResult(
                doi=doi,
                authors=authors,
                venue=clean_text(item.get("venue")),
                abstract=clean_text(item.get("abstract")),
                authority="Semantic Scholar",
                abstract_status="Semantic Scholar abstract" if item.get("abstract") else "",
                source_url=clean_text(item.get("url")),
                score=score,
            )
            best_score = score

    return best if best and best.score >= 0.7 else None


# ----------------------------
# Merge / decision logic
# ----------------------------

def merge_results(current_row: pd.Series, candidates: List[HarvestResult]) -> Dict[str, str]:
    merged = {
        "DOI": clean_text(current_row.get("DOI")),
        "Authors": clean_text(current_row.get("Authors")),
        "Venue": clean_text(current_row.get("Venue")),
        "Abstract": clean_text(current_row.get("Abstract")),
        "Metadata_Status": clean_text(current_row.get("Metadata_Status")),
        "Harvest_Source": "",
        "Abstract_Source_Status": "",
        "Harvest_Source_URL": clean_text(current_row.get("Source_URL")),
    }

    # prefer strongest source with actual data
    doi_candidate = ""
    authors_candidate = ""
    venue_candidate = ""
    abstract_candidate = ""
    source_url_candidate = ""
    authorities = []
    abstract_statuses = []

    for cand in candidates:
        if not cand:
            continue
        authorities.append(cand.authority)
        if cand.abstract_status:
            abstract_statuses.append(cand.abstract_status)
        doi_candidate = choose_better(doi_candidate, cand.doi)
        authors_candidate = choose_better(authors_candidate, cand.authors)
        venue_candidate = choose_better(venue_candidate, cand.venue)
        abstract_candidate = choose_better(abstract_candidate, cand.abstract)
        source_url_candidate = choose_better(source_url_candidate, cand.source_url)

    if doi_candidate:
        merged["DOI"] = doi_candidate
    if authors_candidate:
        merged["Authors"] = authors_candidate
    if venue_candidate:
        merged["Venue"] = venue_candidate
    if abstract_candidate:
        merged["Abstract"] = abstract_candidate
    if source_url_candidate:
        merged["Harvest_Source_URL"] = source_url_candidate

    merged["Harvest_Source"] = "; ".join(unique_preserve_order(authorities))
    merged["Abstract_Source_Status"] = "; ".join(unique_preserve_order(abstract_statuses))

    status_bits = []
    if merged["Harvest_Source"]:
        status_bits.append(f"Authoritative harvest: {merged['Harvest_Source']}")
    if merged["Abstract_Source_Status"]:
        status_bits.append(merged["Abstract_Source_Status"])
    merged["Metadata_Status"] = "; ".join(unique_preserve_order(
        [clean_text(current_row.get("Metadata_Status"))] + status_bits
    ))
    return merged


# ----------------------------
# IO helpers
# ----------------------------

def load_table(path: Path, sheet: str) -> Tuple[pd.DataFrame, str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, sheet_name=sheet), "excel"
    if suffix == ".csv":
        return pd.read_csv(path), "csv"
    raise ValueError(f"Unsupported input format: {path.suffix}")


def save_table(df: pd.DataFrame, input_path: Path, output_path: Path, sheet: str) -> None:
    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        df.to_csv(output_path, index=False)
        return

    # Preserve workbook sheets where possible
    if input_path.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(input_path)
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(sheet)
        for c_idx, col in enumerate(df.columns, start=1):
            ws.cell(1, c_idx).value = col
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx).value = value
        wb.save(output_path)
        return

    # Fallback
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)


# ----------------------------
# Main batch processor
# ----------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Harvest authoritative metadata for OSINT library records.")
    parser.add_argument("--input", required=True, help="Input workbook or CSV")
    parser.add_argument("--output", required=True, help="Output workbook or CSV")
    parser.add_argument("--sheet", default="Cyber_Corpus", help="Worksheet name if using Excel")
    parser.add_argument("--start-id", type=int, help="Start record ID (inclusive)")
    parser.add_argument("--end-id", type=int, help="End record ID (inclusive)")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="Convenience size; informational")
    parser.add_argument("--mailto", default=os.getenv("HARVEST_MAILTO", ""), help="Contact email for polite API usage")
    parser.add_argument("--cache", default=".harvest_cache.json", help="JSON cache file path")
    parser.add_argument("--semantic-scholar", action="store_true", help="Enable Semantic Scholar fallback")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    cache_path = Path(args.cache)

    df, kind = load_table(input_path, args.sheet)

    required_cols = ["ID", "Title"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    for col in ["DOI", "Authors", "Venue", "Abstract", "Metadata_Status", "Source_URL"]:
        if col not in df.columns:
            df[col] = ""

    for col in ["Harvest_Source", "Abstract_Source_Status", "Harvest_Source_URL"]:
        if col not in df.columns:
            df[col] = ""

    client = CachedHTTPClient(cache_path=cache_path, mailto=args.mailto or None)

    mask = pd.Series([True] * len(df))
    if args.start_id is not None:
        mask &= df["ID"].fillna(-1).astype(int) >= args.start_id
    if args.end_id is not None:
        mask &= df["ID"].fillna(-1).astype(int) <= args.end_id

    work_df = df.loc[mask].copy()
    total = len(work_df)

    print(f"Processing {total} records")
    if args.start_id is not None or args.end_id is not None:
        print(f"ID range: {args.start_id} to {args.end_id}")

    for idx, (row_index, row) in enumerate(work_df.iterrows(), start=1):
        rec_id = row.get("ID")
        title = clean_text(row.get("Title"))
        doi = clean_text(row.get("DOI"))

        print(f"[{idx}/{total}] ID={rec_id} | {title[:120]}")

        candidates: List[HarvestResult] = []

        # 1. DOI-resolved lookups first
        if doi:
            crossref_result = crossref_by_doi(client, doi)
            if crossref_result:
                candidates.append(crossref_result)

            openalex_result = openalex_by_doi(client, doi, args.mailto or None)
            if openalex_result:
                candidates.append(openalex_result)

        # 2. Title lookups
        if not candidates:
            crossref_title = crossref_search_title(client, title)
            if crossref_title:
                candidates.append(crossref_title)

            openalex_title = openalex_search_title(client, title, args.mailto or None)
            if openalex_title:
                candidates.append(openalex_title)

        # 3. Optional S2 fallback
        if args.semantic_scholar and not candidates:
            s2 = semantic_scholar_search_title(client, title)
            if s2:
                candidates.append(s2)

        merged = merge_results(row, candidates)

        for col, value in merged.items():
            df.at[row_index, col] = value

    save_table(df, input_path, output_path, args.sheet)

    print(f"Saved: {output_path}")
    print(f"Cache: {cache_path.resolve()}")
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
